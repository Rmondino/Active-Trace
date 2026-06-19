"""AnalisisService — pure Python computation over academic data.

All analysis logic lives here. Services call Repositories to get raw data,
then compute in Python. NO complex SQL, NO raw queries, NO computed columns.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font


class AnalisisService:
    def __init__(
        self,
        calificacion_repo,
        umbral_repo,
        version_padron_repo,
        entrada_padron_repo,
        asignacion_repo=None,
    ):
        self.calificacion_repo = calificacion_repo
        self.umbral_repo = umbral_repo
        self.version_padron_repo = version_padron_repo
        self.entrada_padron_repo = entrada_padron_repo
        self.asignacion_repo = asignacion_repo

    @staticmethod
    def _mask_email(email: str) -> str:
        if "@" in email:
            local, domain = email.split("@", 1)
            if len(local) <= 3:
                prefix = local
            else:
                prefix = local[:3]
            return f"{prefix}...@{domain}"
        if len(email) <= 3:
            return email
        return email[:3] + "..."

    async def _get_activo_padron(self, materia_id, cohorte_id, tenant_id):
        version = await self.version_padron_repo.get_activa(materia_id, cohorte_id)
        if not version:
            return [], set()
        entradas = await self.entrada_padron_repo.get_by_version(version.id)
        return entradas, version

    async def alumnos_atrasados(
        self, materia_id: str, cohorte_id: str, tenant_id: str
    ) -> list[dict]:
        entradas, _ = await self._get_activo_padron(materia_id, cohorte_id, tenant_id)
        if not entradas:
            return []

        todas_calificaciones = await self.calificacion_repo.get_by_materia(materia_id)
        actividades = sorted({
            c.actividad for c in todas_calificaciones
        })
        califs_por_entrada: dict[str, list] = {}
        for c in todas_calificaciones:
            if c.entrada_padron_id not in califs_por_entrada:
                califs_por_entrada[c.entrada_padron_id] = []
            califs_por_entrada[c.entrada_padron_id].append(c)

        umbral = await self.umbral_repo.get_by_materia(materia_id)
        umbral_pct = umbral.umbral_pct if umbral else 60
        valores_aprob = umbral.valores_aprobatorios if umbral else ["Satisfactorio", "Supera lo esperado"]

        result = []
        for ep in entradas:
            califs = califs_por_entrada.get(ep.id, [])
            califs_por_act = {c.actividad: c for c in califs}

            faltantes = [a for a in actividades if a not in califs_por_act]
            baja_nota = []
            for act, c in califs_por_act.items():
                if c.aprobado:
                    continue
                if c.nota_numerica is not None and c.nota_numerica < umbral_pct:
                    baja_nota.append(act)
                elif c.nota_textual is not None and c.nota_textual not in valores_aprob:
                    baja_nota.append(act)
                elif c.nota_numerica is None and c.nota_textual is None and act not in faltantes:
                    baja_nota.append(act)

            aprobadas = sum(1 for c in califs if c.aprobado)
            desaprobadas = sum(1 for c in califs if not c.aprobado)
            sin_nota = len(actividades) - len(califs)

            result.append({
                "alumno": f"{ep.nombre} {ep.apellidos}",
                "entrada_padron_id": ep.id,
                "email_masked": self._mask_email(ep.email),
                "comision": ep.comision or "",
                "es_atrasado": bool(faltantes) or bool(baja_nota),
                "causas": {
                    "faltantes": faltantes,
                    "baja_nota": baja_nota,
                },
                "total_actividades": len(actividades),
                "aprobadas": aprobadas,
                "desaprobadas": desaprobadas,
                "sin_nota": sin_nota,
            })

        return result

    async def ranking_aprobadas(
        self, materia_id: str, tenant_id: str
    ) -> list[dict]:
        todas_calificaciones = await self.calificacion_repo.get_by_materia(materia_id)
        if not todas_calificaciones:
            return []

        entrada_ids = {c.entrada_padron_id for c in todas_calificaciones if c.entrada_padron_id}

        from app.models.entrada_padron import EntradaPadron
        from sqlalchemy import select
        stmt = select(EntradaPadron).where(EntradaPadron.id.in_(entrada_ids))
        result = await self.calificacion_repo.session.execute(stmt)
        entradas_map = {e.id: e for e in result.scalars().all()}

        agrupado: dict[str, dict] = {}
        for c in todas_calificaciones:
            if not c.entrada_padron_id:
                continue
            eid = c.entrada_padron_id
            if eid not in agrupado:
                agrupado[eid] = {"aprobadas": 0, "total": 0}
            agrupado[eid]["total"] += 1
            if c.aprobado:
                agrupado[eid]["aprobadas"] += 1

        ranking = []
        for eid, data in agrupado.items():
            if data["aprobadas"] < 1:
                continue
            ep = entradas_map.get(eid)
            if not ep:
                continue
            total = data["total"]
            porcentaje = (data["aprobadas"] / total * 100) if total else 0.0
            ranking.append({
                "alumno": f"{ep.nombre} {ep.apellidos}",
                "entrada_padron_id": eid,
                "aprobadas": data["aprobadas"],
                "total": total,
                "porcentaje": round(porcentaje, 1),
            })

        ranking.sort(key=lambda r: r["aprobadas"], reverse=True)
        return ranking

    async def reporte_rapido(
        self, materia_id: str, tenant_id: str
    ) -> dict:
        todas_calificaciones = await self.calificacion_repo.get_by_materia(materia_id)
        actividades = sorted({
            c.actividad for c in todas_calificaciones
        })

        entrada_ids = {c.entrada_padron_id for c in todas_calificaciones if c.entrada_padron_id}
        total_alumnos = len(entrada_ids) if entrada_ids else 0

        if not todas_calificaciones:
            return {
                "total_alumnos": 0,
                "total_actividades": 0,
                "alumnos_atrasados": 0,
                "tasa_aprobacion_gral": 0.0,
                "actividades": [],
            }

        total_grades = len(todas_calificaciones)
        aprobadas_count = sum(1 for c in todas_calificaciones if c.aprobado)
        tasa_aprobacion = round(aprobadas_count / total_grades, 2) if total_grades else 0.0

        per_act = []
        for act in actividades:
            act_cals = [c for c in todas_calificaciones if c.actividad == act]
            n_nota_num = [c for c in act_cals if c.nota_numerica is not None]
            aprob_act = sum(1 for c in act_cals if c.aprobado)
            tasa_act = round(aprob_act / len(act_cals), 2) if act_cals else 0.0
            promedio = round(
                sum(float(c.nota_numerica) for c in n_nota_num) / len(n_nota_num), 1
            ) if n_nota_num else None
            per_act.append({
                "nombre": act,
                "tasa_aprobacion": tasa_act,
                "promedio": promedio,
            })

        atrasados_count = 0
        if entrada_ids and actividades:
            califs_por_entrada: dict[str, list] = {}
            for c in todas_calificaciones:
                if c.entrada_padron_id not in califs_por_entrada:
                    califs_por_entrada[c.entrada_padron_id] = []
                califs_por_entrada[c.entrada_padron_id].append(c)

            for eid in entrada_ids:
                califs = califs_por_entrada.get(eid, [])
                califs_por_act = {c.actividad: c for c in califs}
                faltantes = [a for a in actividades if a not in califs_por_act]
                baja_nota = any(not c.aprobado for c in califs)
                if faltantes or baja_nota:
                    atrasados_count += 1

        return {
            "total_alumnos": total_alumnos,
            "total_actividades": len(actividades),
            "alumnos_atrasados": atrasados_count,
            "tasa_aprobacion_gral": tasa_aprobacion,
            "actividades": per_act,
        }

    async def notas_finales(
        self, materia_id: str, cohorte_id: str, tenant_id: str
    ) -> list[dict]:
        entradas, _ = await self._get_activo_padron(materia_id, cohorte_id, tenant_id)
        if not entradas:
            return []

        todas_calificaciones = await self.calificacion_repo.get_by_materia(materia_id)
        califs_por_entrada: dict[str, list] = {}
        for c in todas_calificaciones:
            if c.entrada_padron_id not in califs_por_entrada:
                califs_por_entrada[c.entrada_padron_id] = []
            califs_por_entrada[c.entrada_padron_id].append(c)

        result = []
        for ep in entradas:
            califs = califs_por_entrada.get(ep.id, [])
            actividades = {}
            sum_numerica = 0
            count_numerica = 0
            aprobadas = 0
            for c in califs:
                if c.nota_numerica is not None:
                    actividades[c.actividad] = float(c.nota_numerica)
                    sum_numerica += float(c.nota_numerica)
                    count_numerica += 1
                elif c.nota_textual is not None:
                    actividades[c.actividad] = c.nota_textual
                else:
                    actividades[c.actividad] = None
                if c.aprobado:
                    aprobadas += 1

            promedio = round(sum_numerica / count_numerica, 1) if count_numerica else None

            result.append({
                "alumno": f"{ep.nombre} {ep.apellidos}",
                "entrada_padron_id": ep.id,
                "actividades": actividades,
                "promedio_numerico": promedio,
                "aprobadas": aprobadas,
                "total_actividades": len(califs),
            })

        return result

    async def exportar_sin_corregir(
        self, materia_id: str, tenant_id: str
    ) -> bytes:
        todas_calificaciones = await self.calificacion_repo.get_by_materia(materia_id)

        textual_sin_nota = [
            c for c in todas_calificaciones
            if c.nota_numerica is None
            and c.nota_textual is None
        ]

        entrada_ids = {c.entrada_padron_id for c in textual_sin_nota if c.entrada_padron_id}
        entradas_map = {}
        if entrada_ids:
            from app.models.entrada_padron import EntradaPadron
            from sqlalchemy import select
            stmt = select(EntradaPadron).where(EntradaPadron.id.in_(entrada_ids))
            result = await self.calificacion_repo.session.execute(stmt)
            entradas_map = {e.id: f"{e.nombre} {e.apellidos}" for e in result.scalars().all()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sin Corregir"

        headers = ["Alumno", "Actividad", "Estado"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for i, c in enumerate(textual_sin_nota, 2):
            alumno = entradas_map.get(c.entrada_padron_id, "Desconocido") if c.entrada_padron_id else "Desconocido"
            ws.cell(row=i, column=1, value=alumno)
            ws.cell(row=i, column=2, value=c.actividad)
            ws.cell(row=i, column=3, value="Sin corregir")

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def monitor(
        self, scope: str, filtros: dict, user_id: str, tenant_id: str
    ) -> list[dict]:
        materia_ids = set()

        if scope == "propio" and self.asignacion_repo:
            asignaciones = await self.asignacion_repo.get_by_usuario(tenant_id, user_id)
            materia_ids = {
                a.materia_id for a in asignaciones if a.materia_id
            }
        elif scope == "general":
            from app.repositories.materia_repository import MateriaRepository
            mr = MateriaRepository(session=self.calificacion_repo.session, tenant_id=tenant_id)
            materias = await mr.list(limit=1000)
            materia_ids = {m.id for m in materias}

        materia_filter = filtros.get("materia_id")
        if materia_filter:
            materia_ids = {materia_filter} & materia_ids

        all_results = []
        comision_filter = filtros.get("comision")
        busqueda = filtros.get("busqueda", "").lower()

        for mid in materia_ids:
            todas_calificaciones = await self.calificacion_repo.get_by_materia(mid)
            if not todas_calificaciones:
                continue

            actividades = sorted({c.actividad for c in todas_calificaciones})
            califs_por_entrada: dict[str, list] = {}
            for c in todas_calificaciones:
                if c.entrada_padron_id:
                    if c.entrada_padron_id not in califs_por_entrada:
                        califs_por_entrada[c.entrada_padron_id] = []
                    califs_por_entrada[c.entrada_padron_id].append(c)

            entrada_ids = list(califs_por_entrada.keys())
            if not entrada_ids:
                continue

            from app.models.entrada_padron import EntradaPadron
            from sqlalchemy import select
            stmt = select(EntradaPadron).where(EntradaPadron.id.in_(entrada_ids))
            ep_result = await self.calificacion_repo.session.execute(stmt)
            entradas_map = {e.id: e for e in ep_result.scalars().all()}

            umbral = await self.umbral_repo.get_by_materia(mid)
            umbral_pct = umbral.umbral_pct if umbral else 60
            valores_aprob = umbral.valores_aprobatorios if umbral else ["Satisfactorio", "Supera lo esperado"]

            desde = filtros.get("desde")
            hasta = filtros.get("hasta")

            for eid, califs in califs_por_entrada.items():
                ep = entradas_map.get(eid)
                if not ep:
                    continue

                if comision_filter and ep.comision != comision_filter:
                    continue
                if busqueda:
                    full_name = f"{ep.nombre} {ep.apellidos}".lower()
                    if busqueda not in full_name:
                        continue

                califs_por_act = {c.actividad: c for c in califs}

                if desde or hasta:
                    filtered = []
                    for c in califs:
                        if desde and c.importado_at and c.importado_at.date() < desde:
                            continue
                        if hasta and c.importado_at and c.importado_at.date() > hasta:
                            continue
                        filtered.append(c)
                    califs = filtered

                faltantes = [a for a in actividades if a not in califs_por_act]
                baja_nota = [
                    act for act, c in califs_por_act.items()
                    if not c.aprobado
                ]
                es_atrasado = bool(faltantes) or bool(baja_nota)

                estado_filter = filtros.get("estado")
                if estado_filter == "atrasado" and not es_atrasado:
                    continue
                if estado_filter == "no_atrasado" and es_atrasado:
                    continue

                actividad_filter = filtros.get("actividad")
                if actividad_filter:
                    if actividad_filter not in faltantes and actividad_filter not in baja_nota:
                        continue

                aprobadas = sum(1 for c in califs if c.aprobado)

                all_results.append({
                    "alumno": f"{ep.nombre} {ep.apellidos}",
                    "entrada_padron_id": eid,
                    "email_masked": self._mask_email(ep.email),
                    "comision": ep.comision or "",
                    "es_atrasado": es_atrasado,
                    "causas": {
                        "faltantes": faltantes,
                        "baja_nota": baja_nota,
                    },
                    "total_actividades": len(actividades),
                    "aprobadas": aprobadas,
                    "materia_id": mid,
                })

        return all_results
